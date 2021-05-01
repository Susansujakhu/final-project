""" Messing about with tkinter """
import tkinter as tk
from tkinter import Frame, StringVar, ttk
from tkinter import messagebox
from fastai.basic_train import load_learner
from fastai.vision.image import open_image
import numpy as np
from fastai import *
from fastai.vision import *
# from fastai.vision.all import *

import tkinter.filedialog
from openpyxl import Workbook
from datetime import datetime
from os import name, path
from openpyxl import load_workbook

import sys
import os

LARGE_FONT = ("Verdana", 12)

class Window(tk.Tk):
    """ Main class """
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}

        for frame in (Interface, Predict):
            current_frame = frame(container, self)
            self.frames[frame] = current_frame
            current_frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame(Interface)

    def show_frame(self, cont):
        """ Raises a particular frame, bringing it into view """
        frame = self.frames[cont]
        frame.tkraise()

    def get_page(self, page_class):
        return self.frames[page_class]

class Interface(tk.Frame):
    """ Main frame of program """

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.makeWidgets()
        self.tree()
        menubar = tk.Menu(controller)
        mFile = tk.Menu(menubar, tearoff=False)
        menubar.add_cascade(label="File", menu=mFile)

        # mFile.add_command(label="Open", command=self.checkView)
        # mFile.add_command(label="Save", command=self.check_saveView)
        mFile.add_separator()
        mFile.add_command(label="Exit", command=lambda: quit())

        # Help Section
        mHelp = tk.Menu(menubar, tearoff=False)
        menubar.add_cascade(label="Help", menu=mHelp)

        mHelp.add_command(label="About")
        controller.config(menu=menubar)
        self.controller = controller
        # Hospitall name
        ttk.Label(self,font= ('Arial', 25), text="Knee OA Detection System").place(relx = 0.3, rely =0.01)

        # ******* Buttons ADD and Send ***********
        ttk.Button(self, text= "Save & Send", command= self.add_data).place(
            relx=0.4,rely=0.27)
        ttk.Button(self, text= "Clear", command= self.clear).place(relx=0.5,rely=0.27)
        #***********
        ttk.Button(self, text= "Send", command= lambda : self.goto_predict(controller)).place(relx=0.8,rely=0.9)
        #ttk.Button(self, text= "Send", command= lambda : controller.show_frame(Predict)).place(relx=0.8,rely=0.9)
        
        # ****** Delete record of treeview ********
        ttk.Button(self, text= "Delete", command= self.delete).place(relx=0.1,rely=0.9)
        #****** Search For Treeview***********

        self.search_text = StringVar()
        self.searchBy = StringVar()

        ttk.Label(self, text="Search By").place(relx = 0.65, rely =0.35)

        self.search_by = ttk.Combobox(self, textvariable=self.searchBy, 
                                state='readonly')
        self.search_by['values'] = ('Name', 'Address', 'City')
        self.search_by.current(0)
        self.search_by.place(
             relx = 0.72, rely =0.35, width=100, height=25)

        ttk.Entry(self, font = ('Arial', 12), textvariable = self.search_text).place(
            relx = 0.8, rely =0.35, width=150, height=25)
        self.search_text.trace("w",self.filterSearch)

    def filterSearch(self, *args):
        self.category = self.searchBy.get()
        j = 1
        for i in self.column:
            j = j + 1
            if self.category == i:
                break
        
        ItemsOnTreeview = self.tree.get_children()
        search = self.search_text.get()
        print(search.lower())
        for eachItem in ItemsOnTreeview:
            if search in self.tree.item(eachItem)['values'][j-1].lower():
                #print(self.tree.item(eachItem)['values'][2])
                search_var = self.tree.item(eachItem)['values']
                self.tree.delete(eachItem)

                self.tree.insert("", 0, values=search_var)


    def makeWidgets(self):
        self.id_text = StringVar()
        self.name_text = StringVar()
        self.address_text = StringVar()
        self.city_text = StringVar()
        self.gender_value = StringVar()
        self.age_value = StringVar()
        self.contact_text = StringVar()
        self.blood_value = StringVar()
        
        
        style = entryStyle = labelStyle = buttonStyle = ttk.Style()
        style.configure('.', font=('Helvetica', 12))
        entryStyle.configure('TEntry',foreground = 'green')
        labelStyle.configure('TLable', )
        buttonStyle.configure('TButton',background='#232323', foreground = 'black', borderwidth=1, focusthickness=3, focuscolor='green')
        buttonStyle.map('TButton', background=[('active','green')])

        # Patient Id
        ttk.Label(self, text="Patient Id:").place(relx = 0.05, rely =0.1)
        ttk.Entry(self, font = ('Arial', 12), textvariable = self.id_text, style="TEntry").place(
            relx = 0.11, rely =0.1, width=200, height=25)

        # Name
        ttk.Label(self, text="Full Name:").place(relx = 0.3, rely =0.1)
        ttk.Entry(self, font = ('Arial', 12), textvariable = self.name_text, style="TEntry").place(
            relx = 0.37, rely =0.1, width=250, height=25)

        # Gender
        ttk.Label(self, text="Gender:").place(relx = 0.6, rely =0.1)
       
        # Gender Combobox
        self.gender = ttk.Combobox(self, textvariable=self.gender_value, 
                                state='readonly')
        self.gender['values'] = ('None', 'Male', 'Female')
        self.gender.current()
        self.gender.place(
             relx = 0.65, rely =0.1, width=100, height=25)

        # Age
        ttk.Label(self, text="Age:").place(relx = 0.77, rely =0.1)
        ttk.Entry(self, font = ('Arial', 12), textvariable = self.age_value).place(
            relx = 0.8, rely =0.1, width=100, height=25)
        
        #********************************************************************

        # Blood Group
        ttk.Label(self, text="Blood Group:").place(relx = 0.05, rely =0.18)
        self.blood_group = ttk.Combobox(self, textvariable=self.blood_value, 
                                state='readonly')
        self.blood_group['values'] = ('A+', 'A-', 'B+', 'B-', 'O+', 'O-', 'AB+', 'AB-')
        self.blood_group.current()
        self.blood_group.place(
            relx = 0.13, rely =0.18, width=150, height=25)

        # Contact Number
        ttk.Label(self, text="Contact No.:").place(relx = 0.3, rely =0.18)
        ttk.Entry(self, font = ('Arial', 12), textvariable = self.contact_text, style="TEntry").place(
            relx = 0.37, rely =0.18, width=200, height=25)

        # Address
        ttk.Label(self, text="Address:").place(relx = 0.55, rely =0.18)
        ttk.Entry(self, font = ('Arial', 12), textvariable = self.address_text).place(
            relx = 0.6, rely =0.18, width=200, height=25)
        
        # City
        ttk.Label(self, text="City:").place(relx = 0.77, rely =0.18)
        ttk.Entry(self, font = ('Arial', 12), textvariable = self.city_text).place(
            relx = 0.8, rely =0.18, width=100, height=25)

        # ***********************************
    def clear(self):
        self.id_text.set("")
        self.name_text.set("")
        self.address_text.set("")
        self.city_text.set("")
        self.gender_value.set("")
        self.age_value.set("")
        self.contact_text.set("")
        self.blood_value.set("")

    def tree(self):

        frame1 = Frame(self)
        frame1.pack()
        frame1.place(relheight=0.5, relwidth=0.9, relx = 0.05, rely =0.4)

        tk.Grid.rowconfigure(frame1, 0, weight=1)
        tk.Grid.columnconfigure(frame1, 0, weight=1)

        self.tree = ttk.Treeview(frame1,selectmode="browse", show="headings", height=2)
        self.tree.grid(column=0, row=0, sticky='news')
        ## Adds scrollbars
        wY = ttk.Scrollbar(frame1, orient="vertical", command=self.tree.yview)
        wY.grid(column=1, row=0, sticky='ns')
        wY.config(takefocus=0)

        wX = ttk.Scrollbar(frame1, orient="horizontal", command=self.tree.xview)
        wX.grid(column=0, row=1, sticky='we')
        wX.config(takefocus=0)

        self.tree.configure(xscrollcommand=wX.set, yscrollcommand=wY.set)
        
        #column_index = ["1","2","3", "4", "5", "6", "7", "8", "9", "10"]
        
        self.column = ['Patient ID', 'Name', 'Gender', 'Age', 'Blood Group', 'Contact', 'Address', 'City', 'Date Created', 'Result']
        self.tree["columns"] = ['#'] + self.column

        self.tree.heading('#', text='#')
        self.tree.column('#', minwidth=20, width=30, stretch=False)

        for i in self.column:
            self.tree.column(i, width=100, minwidth=50)
            self.tree.heading(i, text = i)
        
        if path.exists("data.xlsx"):
            book = load_workbook("data.xlsx")
            iterRows = iter(book.active)

            for i, row in enumerate(iterRows, 1):
                if i != 1:
                    rowData = [ cell.value for cell in row ]
                    self.tree.insert('', "end", values=rowData)
            book.close()

    def add_data(self):
        now = datetime.now()
        dt_string = now.strftime("%d-%m-%Y-%H-%M-%S")
        self.tree_rows = len(self.tree.get_children())
        self.data = [self.tree_rows+1, self.id_text.get(), self.name_text.get(), self.gender_value.get(), self.age_value.get(),
        self.blood_value.get(), self.contact_text.get(), self.address_text.get(), self.city_text.get(), dt_string, ""]
        self.tree.insert('', tk.END, values=self.data)
        self.saveView()

    def saveView(self):
        
        if path.exists("data.xlsx"):
            wb = load_workbook('data.xlsx')
            work_sheet = wb.active # Get active sheet
            work_sheet.append(self.data)
            wb.save('data.xlsx')

        else:
            book = Workbook()
            sheet = book.active
            headIndex = 1
            for head in self.tree['column']:
                sheet.cell(row=1, column=headIndex).value = head
                headIndex += 1

            rowIndex = 2
            for row in self.tree.get_children():
                colIndex = 1
                for value in self.tree.item(row)['values']:
                    sheet.cell(row = rowIndex, column = colIndex).value = value
                    colIndex += 1
                rowIndex += 1
            book.save('data'+".xlsx")
            book.close()
        child_id = self.tree.get_children()[-1]
        self.tree.focus(child_id)
        self.tree.selection_set(child_id)
        res = messagebox.askquestion(title = "Save",message = "Data Saved. DO you wnat to send for Prediction?")
        if res == 'yes':
            self.goto_predict(self.controller)
        else:
            messagebox.showwarning('error', 'Something went wrong!')

    def delete(self):
        self.selected_item = self.tree.selection()## get selected item

        for items in self.selected_item:
            self.row = items
        self.row = self.row.replace('I','')
        self.row = int(self.row) + 1
        self.tree.delete(self.selected_item)
        wb = load_workbook('data.xlsx')
        work_sheet = wb.active # Get active sheet
        work_sheet.delete_rows(self.row)
        wb.save('data.xlsx')

            
    def goto_predict(self, cont):
        self.selected_items = self.tree.selection()
        
        for items in self.selected_items:
            self.row = self.tree.item(items)

        self.selected_row = self.row['values']

        self.idd = self.selected_row[1]

        cont.show_frame(Predict)


class Predict(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)

        self.controller = controller
        self.get_row = self.controller.get_page(Interface)
        # Move to Previous Frame
        ttk.Button(self, text= "Page 1", command= lambda : controller.show_frame(Interface)).place(x=200, y=100)

        ttk.Label(self, text="Predictions:").place(relx = 0.5, rely =0.07)


    def tkraise(self):
        self.row = self.get_row.selected_row
        print(self.row)
        tk.Frame.tkraise(self)

        ttk.Label(self, text="Details").place(relx = 0.1, rely =0.07)
        ttk.Label(self, text="ID :").place(relx = 0.05, rely =0.12)
        ttk.Label(self, text="Name :").place(relx = 0.05, rely =0.17)
        ttk.Label(self, text="Gender :").place(relx = 0.05, rely =0.22)
        ttk.Label(self, text="Age :").place(relx = 0.05, rely =0.27)
        
        ttk.Label(self, text=self.row[0]).place(relx = 0.1, rely =0.12)

        # Load Model
        self.x = load_learner('F:\\8thproject\\', 'final.pkl') 
        self.fileName = ""
        # File Explorer Button
        ttk.Button(self, text= "Open File", command= self.fileOpen).place(x=100, y=0)

        # Predict Button
        ttk.Button(self, text= "Predict", 
        command= self.showResult).place(x=100,y=100)

    def fileOpen(self):
        FILE_name = tk.filedialog.askopenfilename(
            initialdir = ".",
            title      = "Open",
            filetypes  = (
                ("Photo files", "*.png"),
                ("All", "*.*")
            )
        )
        if FILE_name:
            self.fileName = FILE_name
    
    def showResult(self):
        if self.fileName == "":
            messagebox.showinfo(title = "Alert",message = "Please Open Any File First")
        else:
            img = open_image(self.fileName)
            predict = self.x.predict(img)
            max_value = max(predict[2],key=lambda x:float(x)) 
            print("KL Grade : "+ str(predict[0]) + " with value " + str(max_value.item()*100))
            result = "KL Grade : "+ str(predict[0]) + " with value " + str(max_value.item()*100)
            ttk.Label(self, text= result).place(relx = 0.3, rely =0.1)
    

def main():
    
    root = Window()
    root.title('xl2web')
    width = root.winfo_screenwidth()
    height = root.winfo_screenheight()
    root.geometry('%dx%d+0+0'% (width, height))
    root.state('zoomed')
   
    root.bind('<Key-Escape>', lambda event: quit())
    root.update()

    return root


if __name__ == '__main__':

    main().mainloop()
    
    