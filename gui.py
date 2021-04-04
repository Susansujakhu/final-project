import tkinter as tk
import tkinter.ttk as ttk
import tkinter.filedialog
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import web


class Window(tk.Frame):
    def __init__(self, parent=None, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)

        self.parent = parent
        self.parent.config(menu=self.makeMenu())
        self.makeWidgets()


    def makeMenu(self):
        mMain  = tk.Menu(self.parent)

        # File Section
        mFile = tk.Menu(mMain, tearoff=False)
        mMain.add_cascade(label="File", menu=mFile)

        mFile.add_command(label="Open", command=self.checkView)
        mFile.add_command(label="Save", command=self.check_saveView)
        mFile.add_separator()
        mFile.add_command(label="Exit", command=lambda: quit())

        # Help Section
        mHelp = tk.Menu(mMain, tearoff=False)
        mMain.add_cascade(label="Help", menu=mHelp)

        mHelp.add_command(label="About")

        return mMain


    def makeWidgets(self):
        self.view = View(self.parent)
        self.view.pack(side="top", fill="both", expand=True)

        # Buttons to select and deselect
        self.flag_selectAll = True
        self.wBtn_selectAll = ttk.Button(
            master  = self.parent,
            text    = "Select all",
            command = self.changeButton
        )
        self.wBtn_selectAll.pack(side="left", pady=10)

        self.wBtn_send = ttk.Button(
            master = self.parent,
            text = "Send",
            command = self.view.sendWeb
        )
        self.wBtn_send.pack(side="left", pady=10)

        self.wBtn_addQueue = ttk.Button(
            master = self.parent,
            text = "Add To Queue",
            command = self.view.add_to_queue
        )
        self.wBtn_addQueue.pack(side="left", pady=10)


        self.wBtn_removeQueue = ttk.Button(
            master = self.parent,
            text = "Remove Queue",
            command = self.view.removeQueue
        )
        self.wBtn_removeQueue.pack(side="left", pady=10)


    def changeButton(self):
        if self.flag_selectAll == True:
            self.view.wTv.selection_set(
                    self.view.wTv.get_children()
                )
            self.wBtn_selectAll["text"] = "Unselect All"
            self.flag_selectAll=False
        else:
            self.view.wTv.selection_remove(
                self.view.wTv.get_children()
                )
            self.wBtn_selectAll["text"] = "Select All"
            self.flag_selectAll=True


    def checkView(self):
        if self.view.emptyView == True:
            self.xlOpen()
        else:
            self.popupAppend()


    def popupAppend(self):
        answer = messagebox.askyesno("Question","Do you want to append?")
        if answer is True:
            self.xlOpen()
        else:
            self.append()


    def check_saveView(self):
        if self.view.emptyView == True:
             messagebox.showinfo(title = "Alert",message = "Please Open Any File First")
        else:
            self.view.saveView()


    def popupAppend_old(self):
        self.popup = tk.Tk()
        self.popup.wm_title("Append")
        label = ttk.Label(self.popup, text="Do you want to Append?")
        label.pack(side="top", fill="x", pady=10)
        B1 = ttk.Button(self.popup, text="Yes", command = lambda:[self.xlOpen(), self.popup.destroy()])
        B1.pack()
        B2 = ttk.Button(self.popup, text="No", command = lambda:[self.append(), self.popup.destroy()])
        B2.pack()


    def append(self):
        self.view.wTv.delete(*self.view.wTv.get_children())
        self.view.j = 0
        self.xlOpen()


    def xlOpen(self):
        FILE_name = tk.filedialog.askopenfilename(
            initialdir = ".",
            title      = "Open",
            filetypes  = (
                ("xlxs files", "*.xlsx"),
                ("All", "*.*")
            )
        )
        if FILE_name:
            self.view.xlLoad(FILE_name)


class View(tk.Frame):
    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)

        self.parent = parent
        self.j=0
        self.emptyView = True
        self.makeWidget()
        self.bindings()


    def makeWidget(self):
        tk.Grid.rowconfigure(self, 0, weight=1)
        tk.Grid.columnconfigure(self, 0, weight=1)

        self.wTv = ttk.Treeview(self,selectmode="extended", show="headings")
        self.wTv.grid(column=0, row=0, sticky='news')
        ## Adds scrollbars
        wY = ttk.Scrollbar(self, orient="vertical", command=self.wTv.yview)
        wY.grid(column=1, row=0, sticky='ns')
        wY.config(takefocus=0)

        wX = ttk.Scrollbar(self, orient="horizontal", command=self.wTv.xview)
        wX.grid(column=0, row=1, sticky='we')
        wX.config(takefocus=0)

        self.wTv.configure(xscrollcommand=wX.set, yscrollcommand=wY.set)


    def bindings(self):
        self.wTv.bind("<Double-Button-1>", self.sendWeb_single)
        self.wTv.bind("q", self.add_to_queue)


    def xlLoad(self, FILE_name):
        book = openpyxl.load_workbook(FILE_name)
        iterRows = iter(book.active)

        heading = [ cell.value for cell in next(iterRows) ]
        self.wTv['column'] = ['#','Queue'] + heading

        self.wTv.heading('#', text='#')
        self.wTv.column('#', minwidth=20, width=30, stretch=False)

        self.wTv.heading('Queue', text='Queue')
        self.wTv.column('Queue', minwidth=40, width=50, stretch=False)
        for col in heading:
            self.wTv.column(col, minwidth=100)
            self.wTv.heading(col, text=col)

        for i, row in enumerate(iterRows, 1):
            if self.j != i:
                i = self.j+1
            rowData = [i] + [""] + [ cell.value for cell in row ]
            self.wTv.insert('', "end", values=rowData)
            self.j=i
        self.emptyView = False
        book.close()


    def saveView(self):
        book = Workbook()
        sheet = book.active
        headIndex = 1
        for head in self.wTv['column']:
            sheet.cell(row=1, column=headIndex).value = head
            headIndex += 1

        rowIndex = 2
        for row in self.wTv.get_children():
            colIndex = 1
            for value in self.wTv.item(row)['values']:
                sheet.cell(row = rowIndex, column = colIndex).value = value
                colIndex += 1
            rowIndex += 1

        now = datetime.now()
        dt_string = now.strftime("%d-%m-%Y-%H-%M-%S")
        book.save(dt_string+".xlsx")
        messagebox.showinfo(title = "Save",message = "File Save Successful")
        book.close()


    def sendWeb_single(self, event):
        item_id = event.widget.focus()
        #print(item_id)
        row = event.widget.item(item_id)
        col = self.wTv['column']
        val = (row['values'])
        maping = dict(zip(col, val))
        web.run(maping)


    def sendWeb(self, event=None):
        selected_items = self.wTv.selection()
        for items in selected_items:
            row = self.wTv.item(items)


    def add_to_queue(self, event=None):
        selected_items = self.wTv.selection()
        for items_id in selected_items:
            self.wTv.set(items_id, '#2', 'Queued')
            item_row = self.wTv.item(items_id)
            val = (item_row['values'])
            self.wTv.item(items_id, values=(val))
            print(items_id)
            print(val)


    def removeQueue(self):
        selected_items = self.wTv.selection()
        for items_id in selected_items:
            self.wTv.set(items_id, '#2', '')


def main():
    root = tk.Tk()
    root.title('xl2web')
    root.geometry('800x400')

    gui = Window(root)
    gui.pack(fill="both", expand=True)
    root.bind('<Key-Escape>', lambda event: quit())
    root.update()

    return root


if __name__ == '__main__':
    main().mainloop()
