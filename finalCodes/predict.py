""" Messing about with tkinter """
import tkinter as tk
from tkinter import Frame, PhotoImage, StringVar, ttk
from tkinter import messagebox
from tkinter import font
#from tkinter import font
from tkinter.constants import DISABLED, RIDGE, X
from tkinter.font import BOLD
# from fastai.basic_train import load_learner
# from fastai.vision.image import open_image

import numpy as np
from fastai import *
from fastai.vision import *

import tkinter.filedialog
import openpyxl
from openpyxl import Workbook
from datetime import datetime
from os import name, path
from openpyxl import load_workbook
from PIL import Image, ImageTk, ImageDraw
import report_generator

import search
import interface

from detecto import core, utils, visualize
from detecto.visualize import show_labeled_image, plot_prediction_grid
from torchvision import transforms
import matplotlib.pyplot as plt
import numpy as np

class Predict(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)


        self.controller = controller
        self.model = core.Model.load('models/localizing_model.pth', ['knee'])
    def tkraise(self):
        style = ttk.Style()
        style.configure('W.TLabel', font=('Times New Roman', 20, BOLD))

        

        # Move to Previous Frame
        #ttk.Button(self, text= "Back", command= lambda : self.controller.show_frame(interface.Interface) if self.back_flag == True else self.controller.show_frame(search.SearchUser)).place(relx = 0, rely =0)
        

        
        
        #print(self.row)
        #print(self.get_data.item_row)
        
        #Deetails Frame
        frame1 = Frame(self, border= 10, relief= RIDGE)
        frame1.pack()
        frame1.place(relheight=1, relwidth=0.25, relx = 0, rely =0)

        #The X-ray Frame
        self.image_frame = Frame(self, border= 5, relief= RIDGE)
        self.image_frame.pack()
        self.image_frame.place(relheight=0.5, relwidth=0.5, relx = 0.35, rely =0.1)
        
        ttk.Label(self.image_frame, text = "Upload X-ray Image", foreground= 'grey', style= 'W.TLabel').place(
            relx = 0.35, rely =0.4)

        #Results Frame
        frame2 = Frame(self, border= 5, relief= RIDGE)
        frame2.pack()
        frame2.place(relheight=1, relwidth=1, relx = 0.25, rely =0.65)

        ttk.Label(self, text="Predictions", style = 'W.TLabel').place(relx = 0.55, rely =0.02)
        # File Explorer Button
        ttk.Button(self, text= "Open Image", cursor="hand2", command= self.fileOpen).place(relx = 0.87, rely = 0.1)

        # Predict Button
        ttk.Button(self, text= "Predict", cursor="hand2", command= self.showResult).place(relx = 0.87, rely = 0.56)
        self.get_data = self.controller.get_page(interface.Interface)
        if self.get_data.from_form == False:
            self.get_data = self.controller.get_page(search.SearchUser)
            self.back_flag = False
        else:
            self.get_data = self.controller.get_page(interface.Interface)
            self.back_flag = True
        # print(self.get_data.selected_id)
        self.selected_id = self.get_data.selected_id
        book = load_workbook("data.xlsx")
        active_book = book.active
        iterRows = iter(book.active)
        for i, row in enumerate(iterRows, 1):
            if i != 1:
                rowData = [ cell.value for cell in row ]
                if str(self.selected_id) == str(rowData[0]):
                    self.row_data = rowData
                    break
        self.column = ['Patient ID', 'Name', 'Gender', 'Age', 'Blood Group', 'Contact', 'Address', 'City', 'Description', 'Image', 'Result', 'Date Created']
        self.col_name = self.column

        #Back Button        
        back_image = Image.open("backBlack.png")
        back_image = back_image.resize((25, 20), Image.ANTIALIAS)
        back = ImageTk.PhotoImage(back_image)
        back_btn = tk.Button(frame1, image = back, 
                                borderwidth = 0,
                                command= lambda : self.controller.show_frame(interface.Interface) if self.back_flag == True else self.controller.show_frame(search.SearchUser), 
                                cursor="hand2")
        back_btn.place(relx = 0.005, rely = 0.005, width=27, height=22)
        back_btn.image = back

        ttk.Label(self, text="Results", style = 'W.TLabel').place(relx = 0.3, rely =0.68)
        tk.Label(self, text= "KL Grade : ", font=('TkTextFont', 10, BOLD)).place(relx = 0.3, rely =0.75)
        tk.Label(self, text= "Accuracy : ",  font=('TkTextFont', 10, BOLD)).place(relx = 0.3, rely =0.8)

        ttk.Label(self, text="Details", style = 'W.TLabel').place(relx = 0.08, rely =0.1)
        m = 0
        row_pos = 0.18
        for each in self.col_name:
            tk.Label(self, text=each +" :", font= ('TkTextFont', 10, BOLD)).place(relx = 0.05, rely = row_pos)
            
            tk.Label(self, text=self.row_data[m], font= ('TkTextFont', 10)).place(relx = 0.13, rely = row_pos + 0.005)

            row_pos = row_pos + 0.05
            m += 1
            if m == 8:
                break

        tk.Frame.tkraise(self)
        

    def fileOpen(self):
        FILE_name = tk.filedialog.askopenfilename(
            initialdir = ".",
            title      = "Open",
            filetypes  = (
                ("Photo files", "*.png"),
                ("All", "*.*")
            )
        )
        if FILE_name:
            self.fileName = FILE_name
            image1 = Image.open(self.fileName)
            predictions = self.model.predict(image1)
            labels, boxes, scores = predictions
            coord = boxes.numpy()[0]
            # show_labeled_image(image1, boxes, labels)

            # test = image1.resize((300, 300), Image.ANTIALIAS)
            original_image = image1
            test = ImageDraw.Draw(original_image)  
            test.rectangle([coord[0], coord[1], coord[2], coord[3]], outline ="red")
            test.text((coord[0],coord[1]), "Knee", fill=(255,0,0))

            original_width, original_height = original_image.size

            original = original_image.resize((int((original_width * 250 )/original_height),250))
            original = ImageTk.PhotoImage(original)
            label1 = tkinter.Label(image=original)

            label1.image = original
            ttk.Label(self.image_frame, image=original).place(relx = 0.2, rely =0.01)
            ttk.Label(self, text= self.fileName).place(relx = 0.26, rely =0.19)
    
            self.cropped_img = image1.crop((coord[0], coord[1], coord[2], coord[3]))
            cropped_image = ImageTk.PhotoImage(self.cropped_img)
            label2 = tkinter.Label(image=cropped_image)
            label2.image = cropped_image
            ttk.Label(self.image_frame, image=cropped_image).place(relx = 0.1, rely =0.01)

    def showResult(self):
        if self.fileName == "":
            messagebox.showinfo(title = "Alert",message = "Please Open Any File First")
        else:
            img = open.image(self.fileName)
            # img = open.image(self.cropped_img)
            # Load Model
            # self.x = load_learner('F:\\8thproject\\', 'final.pkl') 
            self.x = load_learner('models/','MedicalExpert-Iresnet_final.pkl') 
            # self.fileName = ""
            
            predict = self.x.predict(img)

            max_value = max(predict[2],key=lambda x:float(x))
            #print("KL Grade : "+ str(predict[0]) + " with value " + str(max_value.item()*100))
            self.result = "KL Grade : "+ str(predict[0]) + " with value " + str(round(max_value.item()*100, 2))
            grade = str(predict[0])
            accuracy_percent = str(max_value.item()*100)
            ttk.Label(self, text= grade).place(relx = 0.38, rely =0.75)
            ttk.Label(self, text= accuracy_percent).place(relx = 0.38, rely =0.8)

            ttk.Button(self, text= "Save Result", command= self.save_result).place(relx = 0.38, rely = 0.85)
            ttk.Button(self, text= "Download Result", command= self.download).place(relx = 0.5, rely =0.85)

    def save_result(self):
        wb = load_workbook('data.xlsx')
        self.ws = wb.active # Get active sheet
        # print(self.get_data.item_row)
        # row_number = self.get_data.item_row
        # cell = "J"+ str(row_number + 1)
        #  
        row_id = self.row_data[0]


        wb = load_workbook('data.xlsx')
        ws = wb.active # Get active sheet

                # Find Cell Place in Data
        for i in range(2, ws.max_row + 1):
            for j in range(1,2):
                # print(ws.cell(i,j).value)
                if str(row_id) == str(ws.cell(i,j).value):
                    print(ws.cell(i,j).value)
                    cell_place = (ws.cell(i,j))  


        cell_id = cell_place.coordinate

        cell_id_result = cell_id.replace('A','K')
        cell_id_image = cell_id.replace('A','J')
        
        ws[cell_id_result] = self.result
        ws[cell_id_image] = self.fileName

        wb.save('data.xlsx')
        #****** Update Treeview*********
        ItemsOnTreeview = self.get_data.tree.get_children()

        for eachItem in ItemsOnTreeview:
                self.get_data.tree.delete(eachItem)
        # refresh data from xlsl file
        book = load_workbook("data.xlsx")
        iterRows = iter(book.active)
        for i, row in enumerate(iterRows, 1):
            if i != 1:
                rowData = [ cell.value for cell in row ]
                self.get_data.tree.insert('', "end", values=rowData)
        book.close()

        messagebox.showinfo(title = "Success",message = "Result Saved Successful")
    
    def download(self):
        self.save_result()
        report_generator.makePDF(self)