""" Messing about with tkinter """
import tkinter as tk
from tkinter import Frame, PhotoImage, StringVar, ttk
from tkinter import messagebox
from tkinter import font
#from tkinter import font
from tkinter.constants import DISABLED, RIDGE, X
from tkinter.font import BOLD, Font
# from fastai.basic_train import load_learner
# from fastai.vision.image import open_image

import numpy as np
from fastai import *
from fastai.vision import *

import tkinter.filedialog
import openpyxl
from openpyxl import Workbook
from datetime import datetime
from os import name, path
from openpyxl import load_workbook
from PIL import Image, ImageTk
import report_generator
import interface
import predict
import first

LARGE_FONT = ("Verdana", 12)


class SearchUser(tk.Frame):
    """ Main frame of program """

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        ttk.Label(self, text="Patient Details",font= ('Times New Roman', 25, BOLD)).place(relx = 0.35, rely =0.04)

    def changeOnHover(self, button, colorOnHover, colorOnLeave, fgHover, fgLeave):
  
        # background on entering widget
        button.bind("<Enter>", func=lambda e: button.config(background=colorOnHover, foreground = fgHover))

        # background on leaving widget
        button.bind("<Leave>", func=lambda e: button.config(background = colorOnLeave, foreground = fgLeave))
    
    def tkraise(self):
        
        self.trees()
        #Back Button        
        back_image = Image.open("backBlack.png")
        back_image = back_image.resize((25, 20), Image.ANTIALIAS)
        back = ImageTk.PhotoImage(back_image)
        back_btn = tk.Button(self, image = back, 
                                borderwidth = 0,
                                command= lambda : self.controller.show_frame(first.First), 
                                cursor="hand2")
        back_btn.place(relx = 0.005, rely = 0.005, width=27, height=22)
        back_btn.image = back

        sendBtn = tk.Button(self, text= "Send",
                    font=('Segoe UI', 12, BOLD),
                    cursor="hand2",
                    background='darkgreen',
                    foreground='white',
                    activeforeground='white',
                    activebackground='forestgreen',
                    command= lambda : self.goto_predict(self.controller))
        sendBtn.place(relx=0.8,rely=0.9, relwidth=0.09, relheight= 0.05)
        self.changeOnHover(sendBtn, 'forestgreen', 'darkgreen', 'white', 'white')

        # ****** Delete record of treeview ********
        deleteBtn = tk.Button(self, text= "Delete",
                    font=('Segoe UI', 12, BOLD),
                    cursor="hand2",
                    command= self.delete,
                    background='red4',
                    foreground='white',
                    activeforeground='white',
                    activebackground='red')
        deleteBtn.place(relx=0.1,rely=0.9, relwidth=0.09, relheight= 0.05)
        self.changeOnHover(deleteBtn, 'red', 'red4', 'white', 'white')
        
        #****** Search For Treeview***********
        self.searchByX = 0.7
        self.searchByY = 0.055
        self.search_text = StringVar()
        self.searchBy = StringVar()

        ttk.Label(self, text="Search By", font=('Segoe UI', 10)).place(relx = self.searchByX, rely =self.searchByY)

        self.search_by = ttk.Combobox(self, textvariable=self.searchBy, 
                                state='readonly')
        self.search_by['values'] = ('Id', 'Name', 'Address')
        self.search_by.current(0)
        self.search_by.place(relx = self.searchByX + 0.05, rely =self.searchByY - 0.004, width=100, height=25)

        ttk.Entry(self, font = ('Segoe UI', 12), cursor="hand2", textvariable = self.search_text).place(
            relx = self.searchByX + 0.13, rely =self.searchByY - 0.004, width=150, height=25)
        self.search_text.trace("w",self.filterSearch)
        tk.Frame.tkraise(self)
    
    

    def filterSearch(self, *args):
        self.category = self.searchBy.get()
        j = 0
        for i in self.column:
            j = j + 1
            if self.category == i:
                break

        ItemsOnTreeview = self.tree.get_children()
        search = self.search_text.get()
        search = search.lower()
        for eachItem in ItemsOnTreeview:
            if search in self.tree.item(eachItem)['values'][j-1].lower():
                # print(search)
                #print(self.tree.item(eachItem)['values'][2])
                search_var = self.tree.item(eachItem)['values']
                self.tree.delete(eachItem)

                self.tree.insert("", 0, values=search_var)

    def trees(self):

        frame1 = Frame(self)
        frame1.pack()
        frame1.place(relheight=0.8, relwidth=0.9, relx = 0.05, rely =0.1)

        tk.Grid.rowconfigure(frame1, 0, weight=50)
        tk.Grid.columnconfigure(frame1, 0, weight=50)
        style = ttk.Style(frame1)
        style.configure("Treeview.Heading", font=('Segoe UI', 10, BOLD))
        style.configure('Treeview', rowheight=30, font = ('Segoe UI', 10))  #SOLUTION
        self.tree = ttk.Treeview(frame1,selectmode="browse", show="headings", height=2)
        self.tree.grid(column=0, row=0, pady= 2, sticky='news')

        ## Adds scrollbars
        wY = ttk.Scrollbar(frame1, orient="vertical", command=self.tree.yview)
        wY.grid(column=1, row=0, sticky='ns')
        wY.config(takefocus=0)

        wX = ttk.Scrollbar(frame1, orient="horizontal", command=self.tree.xview)
        wX.grid(column=0, row=1, sticky='we')
        wX.config(takefocus=0)

        self.tree.configure(xscrollcommand=wX.set, yscrollcommand=wY.set)
        
        #column_index = ["1","2","3", "4", "5", "6", "7", "8", "9", "10"]
        
        self.column = ['Patient ID', 'Name', 'Gender', 'Age', 'Blood Group', 'Contact', 'Address', 'City', 'Description', 'Image', 'Result', 'Date Created']
        self.tree["columns"] = self.column

        for i in self.column:
            self.tree.column(i, width=150, minwidth=50)
            self.tree.heading(i, text = i)
        
        if path.exists("data.xlsx"):
            self.load_data()
        else:
            book = openpyxl.Workbook()
            sheet = book.active
            headIndex = 1
            for head in self.tree['column']:
                sheet.cell(row=1, column=headIndex).value = head
                headIndex += 1

            rowIndex = 2
            for row in self.tree.get_children():
                colIndex = 1
                for value in self.tree.item(row)['values']:
                    sheet.cell(row = rowIndex, column = colIndex).value = value
                    colIndex += 1
                rowIndex += 1
            book.save('data'+".xlsx")
            book.close()


    def load_data(self):
        book = load_workbook("data.xlsx")
        iterRows = iter(book.active)
        for i, row in enumerate(iterRows, 1):
            if i != 1:
                rowData = [ cell.value for cell in row ]
                self.tree.insert('', "end", values=rowData)
        book.close()

    def add_data(self):
        duplicate_id  = False
        if self.id_text.get() and self.name_text.get() and self.gender_value.get() and self.age_value.get():
            wb = load_workbook('data.xlsx')
            ws = wb.active 
            for i in range(2, ws.max_row + 1):
                for j in range(1,2):
                    if str(self.id_text.get()) == str(ws.cell(i,j).value):
                        duplicate_id = True
            wb.close()
            if duplicate_id == False:
                now = datetime.now()
                dt_string = now.strftime("%d-%m-%Y-%H-%M-%S")
                self.tree_rows = len(self.tree.get_children())

                self.data = [self.id_text.get(), self.name_text.get(), self.gender_value.get(), self.age_value.get(),
                self.blood_value.get(), self.contact_text.get(), self.address_text.get(), self.city_text.get(),
                self.description.get("1.0", 'end-1c') , "", "", dt_string]

                self.tree.insert('', tk.END, values=self.data)
                self.clear()
                self.saveView()
            
            else:
                messagebox.showinfo(title = "Error",message = "Patient ID already exists")
        else:
            self.clear()
            messagebox.showinfo(title = "Error",message = "Please Fill all required fields")


    def saveView(self):
        
        if path.exists("data.xlsx"):
            wb = load_workbook('data.xlsx')
            work_sheet = wb.active # Get active sheet
            work_sheet.append(self.data)
            wb.save('data.xlsx')

        child_id = self.tree.get_children()[-1]
        self.tree.focus(child_id)
        self.tree.selection_set(child_id)
        res = messagebox.askquestion(title = "Save", message = "Data Saved. Do you want to send for Prediction?")
        if res == 'yes':
            self.goto_predict(self.controller)
        # else:
        #     messagebox.showwarning('error', 'Something went wrong!')

    def delete(self):
        self.selected_item = self.tree.selection()## get selected item

        for item in self.tree.selection():
            self.row_item = self.tree.item(item)
        self.selected_row = self.row_item['values']
        to_delete = self.selected_row[0]
        # print(to_delete)

        self.tree.delete(self.selected_item)

        wb = load_workbook('data.xlsx')
        ws = wb.active # Get active sheet

                # Find Cell Place in Data
        for i in range(2, ws.max_row + 1):
            for j in range(1,2):
                if str(to_delete) == str(ws.cell(i,j).value):
                    cell_place = (ws.cell(i,j))  


        delete_cell = cell_place.coordinate

        delete_cell = delete_cell.replace('A','')


        ws.delete_rows(int(delete_cell))
        wb.save('data.xlsx')
        messagebox.showinfo(title = "Success",message = "Data Delete Successful")

            
    def goto_predict(self, controller):

        try:
            for item in self.tree.selection():
                self.row_item = self.tree.item(item)
            self.selected_row = self.row_item['values']
            # print(self.selected_row)
            # self.from_form = False
            self.get_data = self.controller.get_page(interface.Interface)
            self.get_data.from_form = False
            self.selected_id = self.selected_row[0]
            self.controller.show_frame(predict.Predict)
        except:
            messagebox.showinfo(title = "Error",message = "Select any data first")
