""" Messing about with tkinter """
import tkinter as tk
from tkinter import Frame, PhotoImage, StringVar, ttk
from tkinter import messagebox
from tkinter import font
#from tkinter import font
from tkinter.constants import DISABLED, RIDGE, X
from tkinter.font import BOLD
# from fastai.basic_train import load_learner
# from fastai.vision.image import open_image

import numpy as np
from fastai import *
from fastai.vision import *

import tkinter.filedialog
import openpyxl
from openpyxl import Workbook
from datetime import datetime
from os import name, path
from openpyxl import load_workbook
from PIL import Image, ImageTk
import report_generator
import first
import predict
import search
class Interface(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent) #, bg='black')
        parent.configure(background='black')

        menubar = tk.Menu(controller)
        mFile = tk.Menu(menubar, tearoff=False)
        menubar.add_cascade(label="File", menu=mFile)

        # mFile.add_command(label="Open", command=self.checkView)
        # mFile.add_command(label="Save", command=self.check_saveView)
        mFile.add_separator()
        mFile.add_command(label="Exit", command=lambda: quit())

        # Help Section
        mHelp = tk.Menu(menubar, tearoff=False)
        menubar.add_cascade(label="Help", menu=mHelp)

        mHelp.add_command(label="About")
        controller.config(menu=menubar)
        self.controller = controller

        width = self.winfo_screenwidth()
        height = self.winfo_screenheight()
        image1 = Image.open("kneeSimpleRight.png")
        if image1.size != (width, height):
                
            img1_width, img1_height = image1.size
            img1_width = int(img1_width / img1_height) * height
            #img1_height = int(img1_height / img1_width) * width

        image1 = image1.resize((width, height), Image.ANTIALIAS)
        bg = ImageTk.PhotoImage(image1)
        # Show image using label
        bg_label = tk.Label(self, image = bg)
        bg_label.place(x=-1, y=0, relwidth=1, relheight=1)
        bg_label.image = bg
        
        #***********
    def tkraise(self):
        
        
        # Hospitall name
        ttk.Label(self, text="Add Patient", font= ('Algerian', 25, BOLD),foreground = 'white').place(relx = 0.15, rely =0.04)
        self.makeWidgets()
        tk.Frame.tkraise(self)

    def makeWidgets(self):
        self.id_text = StringVar()
        self.name_text = StringVar()
        self.address_text = StringVar()
        self.city_text = StringVar()
        self.gender_value = StringVar()
        self.age_value = StringVar()
        self.contact_text = StringVar()
        self.blood_value = StringVar()
        # self.description = StringVar()
        
        style = entryStyle = labelStyle = buttonStyle = comboBoxStyle = ttk.Style()
        style.configure('.', font=('Helvetica', 12))
        entryStyle.configure('TEntry',foreground = 'green')
        labelStyle.configure('TLabel', background = 'black', foreground = 'white', font = ('Times New Roman', 14, BOLD))
        buttonStyle.configure('TButton',background='#232323', foreground = 'black', borderwidth=1, focusthickness=3, focuscolor='green')
        buttonStyle.map('TButton', background=[('active','green')])
        comboBoxStyle.configure('TCombo', font = ('Arial', 11))

        entryWidth = 320
        entryHeight = 27
        addY = 0.07
        relX = 0.095
        entryRelX = relX + 0.085
        P_relY = 0.15
        N_relY = P_relY + addY
        G_relY = N_relY + addY
        A_relY = G_relY + addY
        C_relY = A_relY + addY
        Add_relY = C_relY + addY
        City_relY = Add_relY + addY
        D_relY = City_relY + addY
        btnX = 0.15
        btnY = D_relY + 0.25

        # Patient Id
        ttk.Label(self, text="*Patient Id:", style='TLabel').place(relx = relX, rely =P_relY)
        ttk.Entry(self, font = ('Arial', 12), textvariable = self.id_text, style="TEntry").place(
            relx = entryRelX, rely =P_relY, width=220, height=entryHeight)

        tk.Button(self, text="Search ID",  
                            command= self.auto_fill,
                            relief= "raised",
                            font=('Arial', 10, BOLD),
                            activeforeground='white',
                            activebackground='green',
                            ).place(relx = relX + 0.26, rely = P_relY, relwidth=0.06, relheight=0.035)   

        # Name
        ttk.Label(self, text="*Full Name:", style='TLabel').place(relx = relX, rely =N_relY)
        ttk.Entry(self, font = ('Arial', 12), textvariable = self.name_text, style="TEntry").place(
            relx = entryRelX, rely =N_relY, width = entryWidth, height=entryHeight)

        # Gender
        ttk.Label(self, text="*Gender:", style='TLabel').place(relx = relX, rely =G_relY)
       
        # Gender Combobox
        self.gender = ttk.Combobox(self, textvariable=self.gender_value, 
                                state='readonly')
        self.gender['values'] = ('Male', 'Female', 'Others')
        #self.gender.current(0)
        self.gender.place(
             relx = entryRelX, rely =G_relY, width=130, height=entryHeight)

        # Age
        ttk.Label(self, text="*Age:", style='TLabel').place(relx = relX, rely = A_relY)
        ttk.Entry(self, font = ('Arial', 12), textvariable = self.age_value).place(
            relx = entryRelX, rely = A_relY, width=90, height=entryHeight)
        
        #********************************************************************

        # Blood Group
        ttk.Label(self, text="Blood Group:", style='TLabel').place(relx = relX + 0.17, rely =A_relY)
        self.blood_group = ttk.Combobox(self, textvariable=self.blood_value, 
                                state='readonly')
        self.blood_group['values'] = ('A+', 'A-', 'B+', 'B-', 'O+', 'O-', 'AB+', 'AB-')
        self.blood_group.current()
        self.blood_group.place(
            relx = relX + 0.2558, rely = A_relY, width=90, height=entryHeight)

        # Contact Number
        ttk.Label(self, text="Contact:", style='TLabel').place(relx = relX, rely = C_relY)
        ttk.Entry(self, font = ('Arial', 12), textvariable = self.contact_text, style="TEntry").place(
            relx = entryRelX, rely = C_relY, width=entryWidth, height=entryHeight)

        # Address
        ttk.Label(self, text="Address:", style='TLabel').place(relx = relX, rely = Add_relY)
        ttk.Entry(self, font = ('Arial', 12), textvariable = self.address_text).place(
            relx = entryRelX, rely = Add_relY, width=entryWidth, height=entryHeight)
        
        # City
        ttk.Label(self, text="City:", style='TLabel').place(relx = relX, rely = City_relY)
        ttk.Entry(self, font = ('Arial', 12), textvariable = self.city_text).place(
            relx = entryRelX, rely = City_relY, width=entryWidth, height=entryHeight)

        # Description
        from tkinter.scrolledtext import ScrolledText
        ttk.Label(self, text="Description:", style='TLabel').place(relx = relX, rely =D_relY)
        self.description = ScrolledText(self, wrap=tk.WORD,
                                      width=42, height=4, border = 2,
                                      font=("Times New Roman", 15))
        self.description.place(relx = relX, rely = D_relY + 0.06)

        # ******* Buttons BACK, ADD and Send *********** 
        #Back Button        
        back_image = Image.open("back3.png")
        back_image = back_image.resize((30, 25), Image.ANTIALIAS)
        back = ImageTk.PhotoImage(back_image)
        back_btn = tk.Button(self, image = back, 
                                borderwidth = 0,
                                command= lambda : self.controller.show_frame(first.First), 
                                background='black',
                                )
        back_btn.place(relx = relX - 0.08, rely = P_relY - 0.12, width=40, height=35)
        back_btn.image = back   

        saveBtn = tk.Button(self, text= "Save",
                            background= 'forestgreen',
                            foreground= 'white',
                            relief='raised',
                            borderwidth = 0,
                            font = ('Times New Roman', 14, BOLD),
                            command= self.add_data)
        saveBtn.place(relx=btnX, rely=btnY, relwidth=0.08, relheight=0.05)
        self.changeOnHover(saveBtn, 'darkgreen', 'forestgreen', 'white', 'white')

        clearBtn = tk.Button(self, text= "Clear",
                            background= 'red', 
                            foreground= 'white',
                            relief='raised',
                            borderwidth = 0,
                            font = ('Times New Roman', 14, BOLD),
                            command= self.clear)
        clearBtn.place(relx=btnX + 0.1, rely = btnY, relwidth=0.08, relheight=0.05)
        self.changeOnHover(clearBtn, 'red4', 'red', 'white', 'white')

        #tk.Button(self, text= "Back", font=('Arial', 10, BOLD),command= lambda : self.controller.show_frame(First)).place(relx = relX - 0.06, rely = P_relY - 0.12, relwidth=0.08, relheight=0.04)
        tk.Button(self, text="Browse All",
                            relief= "raised",
                            borderwidth = 0,
                            font=('Arial', 10, BOLD),
                            command= lambda : self.controller.show_frame(search.SearchUser)
                            ).place(relx = relX + 0.3, rely = P_relY - 0.12, relwidth=0.07, relheight=0.04)  

    def changeOnHover(self, button, colorOnHover, colorOnLeave, fgHover, fgLeave):
  
        # background on entering widget
        button.bind("<Enter>", func=lambda e: button.config(
            background=colorOnHover, foreground = fgHover))

        # background on leaving widget
        button.bind("<Leave>", func=lambda e: button.config(
            background = colorOnLeave, foreground = fgLeave))

    def auto_fill(self):
        book = load_workbook("stored-data.xlsx")
        active_book = book.active
        iterRows = iter(book.active)
        entered_id = self.id_text.get()
        for i, row in enumerate(iterRows, 1):
            if i != 1:
                rowData = [ cell.value for cell in row ]
                if entered_id == rowData[0]:
                    print(rowData[0])
                    self.id_text.set(rowData[0])
                    self.name_text.set(rowData[1])
                    self.gender_value.set(rowData[2])
                    self.age_value.set(rowData[3])
                    self.blood_value.set(rowData[4])
                    self.contact_text.set(rowData[5])
                    self.address_text.set(rowData[6])
                    self.city_text.set(rowData[7])
                    break
            if i == active_book.max_row:
                messagebox.showinfo(title = "Error",message = "Patient ID doesnot exists")

        book.close()


    def clear(self):
        self.id_text.set("")
        self.name_text.set("")
        self.address_text.set("")
        self.city_text.set("")
        self.gender_value.set("")
        self.age_value.set("")
        self.contact_text.set("")
        self.blood_value.set("")
        self.description.delete("1.0","end")

    def add_data(self):
        duplicate_id  = False
        if self.id_text.get() and self.name_text.get() and self.gender_value.get() and self.age_value.get():
            wb = load_workbook('data.xlsx')
            ws = wb.active 
            for i in range(2, ws.max_row + 1):
                for j in range(1,2):
                    if str(self.id_text.get()) == str(ws.cell(i,j).value):
                        duplicate_id = True
            wb.close()
            if duplicate_id == False:
                now = datetime.now()
                dt_string = now.strftime("%d-%m-%Y-%H-%M-%S")

                self.data = [self.id_text.get(), self.name_text.get(), self.gender_value.get(), self.age_value.get(),
                self.blood_value.get(), self.contact_text.get(), self.address_text.get(), self.city_text.get(),
                self.description.get("1.0", 'end-1c') , "", "", dt_string]

                if path.exists("data.xlsx"):
                    wb = load_workbook('data.xlsx')
                    work_sheet = wb.active # Get active sheet
                    work_sheet.append(self.data)
                    wb.save('data.xlsx')

                
                res = messagebox.askquestion(title = "Save",message = "Data Saved. Do you want to send for Prediction?")
                if res == 'yes':
                    self.selected_id = self.id_text.get()
                    self.from_form = True
                    self.controller.show_frame(predict.Predict)
            
            else:
                messagebox.showinfo(title = "Error",message = "Patient ID already exists")
        else:
            # self.clear()
            messagebox.showinfo(title = "Error",message = "Please Fill all required fields")